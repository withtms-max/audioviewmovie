import pandas as pd
import requests
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import os

TMDB_API_KEY = "7f0810596150fed0f54797282c113e33"
INPUT_FILE = "오디오뷰-_목록.xlsx"
OUTPUT_FILE = "영화_정보_완성.xlsx"

print("엑셀 파일 읽는 중...", flush=True)
# 1. Read input Excel
try:
    df = pd.read_excel(INPUT_FILE)
except Exception as e:
    print(f"Failed to read excel: {e}")
    exit(1)

# The movie titles are in 'Unnamed: 1'
movie_titles = df['Unnamed: 1'].dropna().tolist()

# Clean up headers if any
clean_titles = []
for t in movie_titles:
    t_str = str(t).strip()
    if t_str and t_str not in ['영화명', '영화명 ']:
        clean_titles.append(t_str)

print(f"총 {len(clean_titles)} 편의 영화가 확인되었습니다.", flush=True)

# 2. Setup output Workbook
wb = Workbook()
ws = wb.active
ws.title = "Movie Data"

headers = [
    "포스터 이미지", "영화 제목", "장르", "제작국가", "감독", 
    "러닝타임", "개봉일", "등급", "줄거리", "검색 키워드", "포스터 URL"
]
ws.append(headers)

# Set column widths
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 50
ws.column_dimensions['J'].width = 30
ws.column_dimensions['K'].width = 30

def get_movie_data(title):
    search_url = "https://api.themoviedb.org/3/search/movie"
    try:
        res = requests.get(search_url, params={
            "api_key": TMDB_API_KEY,
            "query": title,
            "language": "ko-KR"
        }, timeout=10)
        res.raise_for_status()
        results = res.json().get('results', [])
        if not results:
            # try finding movie without language param
            res = requests.get(search_url, params={
                "api_key": TMDB_API_KEY,
                "query": title
            }, timeout=10)
            results = res.json().get('results', [])
            if not results:
                return None
        
        movie_id = results[0]['id']
        
        # Get details
        details_url = f"https://api.themoviedb.org/3/movie/{movie_id}?api_key={TMDB_API_KEY}&language=ko-KR&append_to_response=credits,release_dates,keywords"
        detail_res = requests.get(details_url, timeout=10)
        detail_res.raise_for_status()
        data = detail_res.json()
        
        # Parse data
        genres = ", ".join([g['name'] for g in data.get('genres', [])])
        countries = ", ".join([c['name'] for c in data.get('production_countries', [])])
        
        director = ""
        credits = data.get('credits', {})
        for crew in credits.get('crew', []):
            if crew.get('job') == 'Director':
                director = crew.get('name')
                break
                
        runtime = f"{data.get('runtime', 0)}분" if data.get('runtime') else ""
        release_date = data.get('release_date', "")
        
        rating = ""
        release_dates = data.get('release_dates', {}).get('results', [])
        for rd in release_dates:
            if rd.get('iso_3166_1') == 'KR': # Korean rating
                for r in rd.get('release_dates', []):
                    if r.get('certification'):
                        rating = r.get('certification')
                        break
                break
                
        if not rating: # fallback to US rating
            for rd in release_dates:
                if rd.get('iso_3166_1') == 'US': 
                    for r in rd.get('release_dates', []):
                        if r.get('certification'):
                            rating = r.get('certification')
                            break
                    break
                    
        plot = data.get('overview', "")
        
        kr_keywords = data.get('keywords', {}).get('keywords', [])
        keywords = ", ".join([k['name'] for k in kr_keywords])
        
        poster_path = data.get('poster_path')
        poster_url = f"https://image.tmdb.org/t/p/w500{poster_path}" if poster_path else ""
        
        return {
            "title": data.get('title', title),
            "genres": genres,
            "countries": countries,
            "director": director,
            "runtime": runtime,
            "release_date": release_date,
            "rating": rating,
            "plot": plot,
            "keywords": keywords,
            "poster_url": poster_url
        }
    except Exception as e:
        print(f"Error fetching {title}: {e}", flush=True)
        return None

row_idx = 2
for i, title in enumerate(clean_titles, 1):
    print(f"[{i}/{len(clean_titles)}] 데이터 수집 중: {title}", flush=True)
    data = get_movie_data(title)
    
    if data:
        ws.cell(row=row_idx, column=2, value=data['title'])
        ws.cell(row=row_idx, column=3, value=data['genres'])
        ws.cell(row=row_idx, column=4, value=data['countries'])
        ws.cell(row=row_idx, column=5, value=data['director'])
        ws.cell(row=row_idx, column=6, value=data['runtime'])
        ws.cell(row=row_idx, column=7, value=data['release_date'])
        ws.cell(row=row_idx, column=8, value=data['rating'])
        ws.cell(row=row_idx, column=9, value=data['plot'])
        ws.cell(row=row_idx, column=10, value=data['keywords'])
        ws.cell(row=row_idx, column=11, value=data['poster_url'])
        
        if data['poster_url']:
            try:
                img_res = requests.get(data['poster_url'], timeout=10)
                if img_res.status_code == 200:
                    os.makedirs("temp_posters", exist_ok=True)
                    img_path = f"temp_posters/poster_{row_idx}.jpg"
                    with open(img_path, "wb") as f:
                        f.write(img_res.content)
                    
                    img = OpenpyxlImage(img_path)
                    img.width, img.height = 100, 150 
                    
                    ws.row_dimensions[row_idx].height = 115
                    ws.add_image(img, f"A{row_idx}")
            except Exception as e:
                print(f"Image load error for {title}: {e}", flush=True)
    else:
        ws.cell(row=row_idx, column=2, value=title)
        ws.cell(row=row_idx, column=9, value="API 정보 없음")
        
    row_idx += 1
    
    # Save periodically to prevent data loss
    if i % 50 == 0:
        wb.save(OUTPUT_FILE)
        print(f"중간 저장 완료 ({i}/{len(clean_titles)})", flush=True)
        
    time.sleep(0.2) 

# Final save
wb.save(OUTPUT_FILE)
print(f"\n✅ 완료되었습니다. 결과물이 저장되었습니다: {OUTPUT_FILE}", flush=True)
