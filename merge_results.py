import pandas as pd
from openpyxl import load_workbook
import os

FILE1 = "영화_정보_완성.xlsx"
FILE2 = "영화_정보_완성_이어서.xlsx"
FINAL_FILE = "영화_정보_최종_합본.xlsx"

print("파일 병합 중...")

# First 200 movies
wb1 = load_workbook(FILE1)
ws1 = wb1.active

# Rest of the movies
wb2 = load_workbook(FILE2)
ws2 = wb2.active

# Iterate through ws2 (skipping header) and add to ws1
for row in ws2.iter_rows(min_row=2):
    # Copy values
    new_row = [cell.value for cell in row]
    ws1.append(new_row)
    
    # Copy images
    # This is tricky because openpyxl doesn't have a simple 'move image'
    # But since we have the temp_posters folder and we know the indices, 
    # we can re-add them if needed, but easier to just let user know 
    # the second file has the rest.
    # Actually, let's try to just append the rows first. 
    # If the user wants images in one file, I might need to re-run the re-adding logic.

# A better way to merge with images:
from openpyxl.drawing.image import Image as OpenpyxlImage

# Since images are lost in pandas/simple merge, let's do it properly via openpyxl
# We'll create a new workbook and add everything.

print("최종 합본 생성 중 (이미지 포함)...")
from openpyxl import Workbook
wb_final = Workbook()
ws_final = wb_final.active
ws_final.title = "Movie Data"

headers = [
    "포스터 이미지", "영화 제목", "장르", "제작국가", "감독", 
    "러닝타임", "개봉일", "등급", "줄거리", "검색 키워드", "포스터 URL"
]
ws_final.append(headers)

# Set widths
col_widths = [16, 25, 20, 15, 15, 10, 12, 10, 50, 30, 30]
for i, width in enumerate(col_widths, 1):
    ws_final.column_dimensions[chr(64+i)].width = width

def copy_data_with_images(src_ws, start_row_idx, resume_prefix=""):
    current_final_row = ws_final.max_row + 1
    
    # Copy text data
    for row in src_ws.iter_rows(min_row=2):
        data = [cell.value for cell in row]
        ws_final.append(data)
        ws_final.row_dimensions[current_final_row].height = 115
        current_final_row += 1

    # Copy images
    for img in src_ws._images:
        # Re-anchor the image to the new row in final ws
        # img.anchor.row is 0-indexed in some contexts? Let's check.
        row_num = img.anchor._from.row + 1 # original row
        new_row_num = (row_num - 2) + start_row_idx
        ws_final.add_image(img, f"A{new_row_num}")

# Add first file
copy_data_with_images(ws1, 2)
# Add second file
copy_data_with_images(ws2, 202) # First file had header + 200 rows = total 201. Next row is 202.

wb_final.save(FINAL_FILE)
print(f"✅ 병합 완료: {FINAL_FILE}")
