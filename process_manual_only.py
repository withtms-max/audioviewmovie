import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import requests
import time

# Configuration
MANUAL_DATA_FILE = "api없는것들.xlsx"
TEMPLATE_FILE = "영화정보_입력양식.xlsx"
OUTPUT_FILE = "추가영화_정리_완료.xlsx"

print(f"'{MANUAL_DATA_FILE}' 파일 처리 시작 (총 41편)...")

# 1. Load Manual Data
if not os.path.exists(MANUAL_DATA_FILE):
    print(f"Error: {MANUAL_DATA_FILE} 파일을 찾을 수 없습니다.")
    exit()

df = pd.read_excel(MANUAL_DATA_FILE)
print(f"불러온 데이터 수: {len(df)}개")

# 2. Excel Operations
wb_template = load_workbook(TEMPLATE_FILE)
ws_template = wb_template.active

# Clear content below header (just in case)
for row in ws_template.iter_rows(min_row=2, max_row=ws_template.max_row):
    for cell in row:
        cell.value = None

# Process each row
for idx, row in df.iterrows():
    row_num = idx + 2
    
    # Mapping based on typical structure preserved in manual file
    # Index-based mapping is safer given encoding in console
    title = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
    genres = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
    country = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""
    director = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
    runtime = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""
    release_date = str(row.iloc[5]) if pd.notna(row.iloc[5]) else ""
    rating = str(row.iloc[6]) if pd.notna(row.iloc[6]) else ""
    plot = str(row.iloc[7]) if pd.notna(row.iloc[7]) else ""
    keywords = str(row.iloc[8]) if pd.notna(row.iloc[8]) else ""
    poster_url = str(row.iloc[9]) if pd.notna(row.iloc[9]) else ""

    print(f"[{idx+1}/{len(df)}] '{title}' 양식 변환 중...")

    # Fill template
    ws_template.cell(row=row_num, column=1, value=title)
    ws_template.cell(row=row_num, column=2, value=genres)
    ws_template.cell(row=row_num, column=3, value=country)
    ws_template.cell(row=row_num, column=4, value=director)
    ws_template.cell(row=row_num, column=5, value=runtime)
    ws_template.cell(row=row_num, column=6, value=release_date)
    ws_template.cell(row=row_num, column=7, value=rating)
    ws_template.cell(row=row_num, column=8, value=plot)
    ws_template.cell(row=row_num, column=9, value=keywords)
    ws_template.cell(row=row_num, column=10, value=poster_url)

    # Image (Optional: try to embed if URL exists)
    if poster_url.startswith("http"):
        try:
            img_res = requests.get(poster_url, timeout=5)
            if img_res.status_code == 200:
                img_path = f"temp_posters/manual_{row_num}.jpg"
                if not os.path.exists("temp_posters"): os.makedirs("temp_posters")
                with open(img_path, "wb") as f:
                    f.write(img_res.content)
                img = OpenpyxlImage(img_path)
                img.width, img.height = 100, 150
                ws_template.row_dimensions[row_num].height = 115
                ws_template.add_image(img, f"K{row_num}")
        except:
            pass

# Set column widths
ws_template.column_dimensions['A'].width = 25
ws_template.column_dimensions['I'].width = 50
ws_template.column_dimensions['K'].width = 16

# Save
wb_template.save(OUTPUT_FILE)
print(f"\n✅ 완료! '{OUTPUT_FILE}' 파일이 생성되었습니다.")
