import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import requests
import time

# Configuration
INPUT_DATA_FILE = "결과물/영화_정보_완성.xlsx"
MANUAL_DATA_FILE = "api없는것들.xlsx"
TEMPLATE_FILE = "영화정보_입력양식.xlsx"
TMDB_API_KEY = "7f0810596150fed0f54797282c113e33"

print("데이터 로딩 시작...")

# 1. Load Primary Data
df = pd.read_excel(INPUT_DATA_FILE)
df = df.drop_duplicates(subset=['영화 제목'], keep='first')

# 2. Load Manual Data if exists
manual_df = None
if os.path.exists(MANUAL_DATA_FILE):
    print(f"수동 입력 데이터({MANUAL_DATA_FILE}) 로딩 중...")
    manual_df = pd.read_excel(MANUAL_DATA_FILE)
    # Map manual columns to our internal format if needed
    # Manual columns: ['영화 명', '장르', '제작국가', '감독', '러닝타임', '개봉일', '등급', '줄거리', '검색 키워드', '포스터 URL']

# 3. Identify "No Info" rows and enrich with Manual Data
def enrich_row(row):
    title = str(row['영화 제목']).strip()
    has_api_info = str(row['줄거리']) != "API 정보 없음" and pd.notna(row['줄거리'])
    
    if not has_api_info and manual_df is not None:
        # Try to find in manual data
        match = manual_df[manual_df.iloc[:, 0].astype(str).str.strip() == title]
        if not match.empty:
            m_row = match.iloc[0]
            row['장르'] = m_row.iloc[1]
            row['제작국가'] = m_row.iloc[2]
            row['감독'] = m_row.iloc[3]
            row['러닝타임'] = m_row.iloc[4]
            row['개봉일'] = m_row.iloc[5]
            row['등급'] = m_row.iloc[6]
            row['줄거리'] = m_row.iloc[7]
            row['검색 키워드'] = m_row.iloc[8]
            row['포스터 URL'] = m_row.iloc[9]
            row['OTT 플랫폼'] = m_row.iloc[10] if len(m_row) > 10 else ""
            row['has_info'] = True # Now it has manual info
            return row
            
    row['has_info'] = has_api_info
    return row

df = df.apply(enrich_row, axis=1)

# 4. Sort: Info first, then No Info
df_info = df[df['has_info']].copy()
df_no_info = df[~df['has_info']].copy()
df_sorted = pd.concat([df_info, df_no_info], ignore_index=True)

# 5. Enrichment Logic for Cast (Fallback if missing)
def get_cast_only(title, orig_cast, row_has_api):
    # If orig_cast already exists and is valid, use it
    if pd.notna(orig_cast) and str(orig_cast).strip() and str(orig_cast) != "nan":
        return str(orig_cast)
        
    # Fetch cast if missing
    cast_names = []
    if row_has_api:
        try:
            search_url = "https://api.themoviedb.org/3/search/movie"
            res = requests.get(search_url, params={"api_key": TMDB_API_KEY, "query": title, "language": "ko-KR"}, timeout=10)
            results = res.json().get('results', [])
            if results:
                movie_id = results[0]['id']
                cred_res = requests.get(f"https://api.themoviedb.org/3/movie/{movie_id}?api_key={TMDB_API_KEY}&language=ko-KR&append_to_response=credits", timeout=10)
                cast_names = [c['name'] for c in cred_res.json().get('credits', {}).get('cast', [])[:10]]
        except:
            pass
            
    return ", ".join(cast_names)

print("최종 양식 파일 생성 중 (수동 데이터 통합 및 키워드 보강)...")

wb_template = load_workbook(TEMPLATE_FILE)
ws_template = wb_template.active

# Clear
for row in ws_template.iter_rows(min_row=2, max_row=ws_template.max_row):
    for cell in row:
        cell.value = None

for idx, row_data in df_sorted.iterrows():
    row_num = idx + 2
    title = str(row_data['영화 제목'])
    
    # We use a flag to check if it's strictly API data for cast enrichment
    # But since manual data can also be enriched, we'll try it if info is present
    orig_genres = row_data.get('장르', '')
    orig_cast = row_data.get('등장인물', row_data.get('검색 키워드', ''))
    director = row_data.get('감독', '')
    
    print(f"[{idx+1}/{len(df_sorted)}] '{title}' 처리 중...", flush=True)
    
    final_cast = get_cast_only(title, orig_cast, row_data['has_info'])
    
    ws_template.cell(row=row_num, column=1, value=title)
    ws_template.cell(row=row_num, column=2, value=orig_genres)
    ws_template.cell(row=row_num, column=3, value=row_data.get('제작국가', ''))
    ws_template.cell(row=row_num, column=4, value=director)
    ws_template.cell(row=row_num, column=5, value=row_data.get('러닝타임', ''))
    ws_template.cell(row=row_num, column=6, value=row_data.get('개봉일', ''))
    ws_template.cell(row=row_num, column=7, value=row_data.get('등급', ''))
    ws_template.cell(row=row_num, column=8, value=row_data.get('줄거리', ''))
    ws_template.cell(row=row_num, column=9, value=final_cast)
    ws_template.cell(row=row_num, column=10, value=row_data.get('포스터 URL', ''))
    ws_template.cell(row=row_num, column=12, value=row_data.get('OTT 플랫폼', ''))
    
    # Image
    poster_url = str(row_data.get('포스터 URL', ''))
    if row_data['has_info'] and poster_url.startswith("http"):
        try:
            img_res = requests.get(poster_url, timeout=10)
            if img_res.status_code == 200:
                os.makedirs("결과물/temp_posters", exist_ok=True)
                img_path = f"결과물/temp_posters/final_merged_{row_num}.jpg"
                with open(img_path, "wb") as f:
                    f.write(img_res.content)
                img = OpenpyxlImage(img_path)
                img.width, img.height = 100, 150
                ws_template.row_dimensions[row_num].height = 115
                ws_template.add_image(img, f"K{row_num}")
        except:
            pass

ws_template.column_dimensions['A'].width = 25
ws_template.column_dimensions['I'].width = 60
ws_template.column_dimensions['K'].width = 16
ws_template.column_dimensions['L'].width = 25

FINAL_OUTPUT = f"결과물/영화정보_입력양식_통합본_{int(time.time())}.xlsx"
wb_template.save(FINAL_OUTPUT)

print(f"\n✅ 통합 완료! 저장된 파일: {FINAL_OUTPUT}")
print(f"정보 있음 (API + 수동): {len(df_info)}")
print(f"정보 없음: {len(df_no_info)}")
