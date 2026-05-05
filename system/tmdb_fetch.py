import pandas as pd
import requests
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
import os

TMDB_API_KEY = "7f0810596150fed0f54797282c113e33"
INPUT_FILE_EXCEL = "오디오뷰-_목록.xlsx"
INPUT_FILE_TXT = "영화목록.txt"
OUTPUT_FILE = "결과물/영화_정보_완성.xlsx"

print("영화 목록 읽는 중...", flush=True)
# 1. Read input
movie_titles = []
if os.path.exists(INPUT_FILE_TXT):
    print(f"[{INPUT_FILE_TXT}] 텍스트 파일에서 읽어옵니다...", flush=True)
    try:
        with open(INPUT_FILE_TXT, 'r', encoding='utf-8') as f:
            movie_titles = [line.strip() for line in f if line.strip()]
    except Exception as e:
        print(f"Failed to read txt: {e}")
        exit(1)
elif os.path.exists(INPUT_FILE_EXCEL):
    print(f"[{INPUT_FILE_EXCEL}] 엑셀 파일에서 읽어옵니다...", flush=True)
    try:
        df = pd.read_excel(INPUT_FILE_EXCEL)
        # B열(Unnamed: 1)에 데이터가 있는지 확인, 없으면 A열(첫 번째 열) 사용
        if 'Unnamed: 1' in df.columns:
            movie_titles = df['Unnamed: 1'].dropna().tolist()
        else:
            movie_titles = df.iloc[:, 0].dropna().tolist()
    except Exception as e:
        print(f"Failed to read excel: {e}")
        exit(1)
else:
    print(f"입력 파일을 찾을 수 없습니다. ({INPUT_FILE_TXT} 또는 {INPUT_FILE_EXCEL}가 필요합니다.)")
    exit(1)

# Clean up headers if any
clean_titles = []
for t in movie_titles:
    t_str = str(t).strip()
    if t_str and t_str not in ['영화명', '영화명 ']:
        clean_titles.append(t_str)

print(f"총 {len(clean_titles)} 편의 영화가 확인되었습니다.", flush=True)

# 2. Setup output Workbook
wb = Workbook()
ws = wb.active
ws.title = "Movie Data"

headers = [
    "포스터 이미지", "영화 제목", "장르", "제작국가", "감독", 
    "러닝타임", "개봉일", "등급", "줄거리", "등장인물", "포스터 URL", "OTT 플랫폼"
]
ws.append(headers)

# Set column widths
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 50
ws.column_dimensions['J'].width = 30
ws.column_dimensions['K'].width = 30
ws.column_dimensions['L'].width = 25

def get_movie_data(title):
    search_url = "https://api.themoviedb.org/3/search/movie"
    try:
        res = requests.get(search_url, params={
            "api_key": TMDB_API_KEY,
            "query": title,
            "language": "ko-KR"
        }, timeout=10)
        res.raise_for_status()
        results = res.json().get('results', [])
        if not results:
            # try finding movie without language param
            res = requests.get(search_url, params={
                "api_key": TMDB_API_KEY,
                "query": title
            }, timeout=10)
            results = res.json().get('results', [])
            if not results:
                return None
        
        movie_id = results[0]['id']
        
        # Get details
        details_url = f"https://api.themoviedb.org/3/movie/{movie_id}?api_key={TMDB_API_KEY}&language=ko-KR&append_to_response=credits,release_dates,keywords,watch/providers"
        detail_res = requests.get(details_url, timeout=10)
        detail_res.raise_for_status()
        data = detail_res.json()
        
        # Parse data
        genres = ", ".join([g['name'] for g in data.get('genres', [])])
        countries = ", ".join([c['name'] for c in data.get('production_countries', [])])
        
        director = ""
        cast_names = []
        credits = data.get('credits', {})
        for crew in credits.get('crew', []):
            if crew.get('job') == 'Director':
                director = crew.get('name')
                break
        
        for cast in credits.get('cast', [])[:10]:
            cast_names.append(cast.get('name'))
        cast_str = ", ".join(cast_names)
                
        runtime = f"{data.get('runtime', 0)}분" if data.get('runtime') else ""
        release_date = data.get('release_date', "")
        
        rating = ""
        release_dates = data.get('release_dates', {}).get('results', [])
        for rd in release_dates:
            if rd.get('iso_3166_1') == 'KR': # Korean rating
                for r in rd.get('release_dates', []):
                    if r.get('certification'):
                        rating = r.get('certification')
                        break
                break
                
        if not rating: # fallback to US rating
            for rd in release_dates:
                if rd.get('iso_3166_1') == 'US': 
                    for r in rd.get('release_dates', []):
                        if r.get('certification'):
                            rating = r.get('certification')
                            break
                    break
                    
        plot = data.get('overview', "")
        
        # 키워드 대신 등장인물 사용
        
        poster_path = data.get('poster_path')
        poster_url = f"https://image.tmdb.org/t/p/w500{poster_path}" if poster_path else ""
        
        # Get OTT Providers
        ott_str = ""
        providers = data.get('watch/providers', {}).get('results', {}).get('KR', {})
        flatrate = providers.get('flatrate', [])
        if flatrate:
            ott_list = [p.get('provider_name') for p in flatrate]
            ott_str = ", ".join(ott_list)
        
        return {
            "title": data.get('title', title),
            "genres": genres,
            "countries": countries,
            "director": director,
            "runtime": runtime,
            "release_date": release_date,
            "rating": rating,
            "plot": plot,
            "cast": cast_str,
            "poster_url": poster_url,
            "ott": ott_str
        }
    except Exception as e:
        print(f"Error fetching {title}: {e}", flush=True)
        return None

row_idx = 2
for i, title in enumerate(clean_titles, 1):
    print(f"[{i}/{len(clean_titles)}] 데이터 수집 중: {title}", flush=True)
    data = get_movie_data(title)
    
    if data:
        ws.cell(row=row_idx, column=2, value=data['title'])
        ws.cell(row=row_idx, column=3, value=data['genres'])
        ws.cell(row=row_idx, column=4, value=data['countries'])
        ws.cell(row=row_idx, column=5, value=data['director'])
        ws.cell(row=row_idx, column=6, value=data['runtime'])
        ws.cell(row=row_idx, column=7, value=data['release_date'])
        ws.cell(row=row_idx, column=8, value=data['rating'])
        ws.cell(row=row_idx, column=9, value=data['plot'])
        ws.cell(row=row_idx, column=10, value=data['cast'])
        ws.cell(row=row_idx, column=11, value=data['poster_url'])
        ws.cell(row=row_idx, column=12, value=data['ott'])
        
        if data['poster_url']:
            try:
                img_res = requests.get(data['poster_url'], timeout=10)
                if img_res.status_code == 200:
                    os.makedirs("결과물/temp_posters", exist_ok=True)
                    img_path = f"결과물/temp_posters/poster_{row_idx}.jpg"
                    with open(img_path, "wb") as f:
                        f.write(img_res.content)
                    
                    img = OpenpyxlImage(img_path)
                    img.width, img.height = 100, 150 
                    
                    ws.row_dimensions[row_idx].height = 115
                    ws.add_image(img, f"A{row_idx}")
            except Exception as e:
                print(f"Image load error for {title}: {e}", flush=True)
    else:
        ws.cell(row=row_idx, column=2, value=title)
        ws.cell(row=row_idx, column=9, value="API 정보 없음")
        
    row_idx += 1
    
    # Save periodically to prevent data loss
    if i % 50 == 0:
        wb.save(OUTPUT_FILE)
        print(f"중간 저장 완료 ({i}/{len(clean_titles)})", flush=True)
        
    time.sleep(0.2) 

# Final save
wb.save(OUTPUT_FILE)
print(f"\n✅ 완료되었습니다. 결과물이 저장되었습니다: {OUTPUT_FILE}", flush=True)
